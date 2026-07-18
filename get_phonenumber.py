import os
import json
import requests
from openpyxl import Workbook, load_workbook

from auth import get_headers

headers = get_headers()

main_api1 = "https://pw-api-gate.penpencil.co/v3/"
comm_url = "community/posts/v2?channelId="

EXCEL_FILE = "details.xlsx"
ID_FILE = "user_id.txt"


def write_unique(post_list):
    user_id = str(post_list[0]).strip()

    if not os.path.exists(ID_FILE):
        open(ID_FILE, "w").close()

    with open(ID_FILE, "r", encoding="utf-8") as f:
        if user_id in {line.strip() for line in f}:
            return

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append([
            "User ID",
            "Name",
            "Mobile No",
            "Profile Pic"
        ])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(post_list)
    wb.save(EXCEL_FILE)
    wb.close()

    with open(ID_FILE, "a", encoding="utf-8") as f:
        f.write(user_id + "\n")


def unfollow(user):
    url = main_api1 + "community/followers/"

    payload = {
        "follow": False
    }

    requests.post(
        url + user,
        headers=get_headers(),
        json=payload,
        timeout=30
    )


def follow(user_id):
    url = main_api1 + "community/followers/"

    payload = {
        "follow": True
    }

    response = requests.post(
        url + user_id,
        headers=get_headers(),
        json=payload,
        timeout=30
    )

    try:
        json_dict = response.json()
        print("Already followed:", not json_dict.get("success", False))
    except Exception:
        print(response.text)


def save(isunfollow=False):
    url = (
        main_api1
        + "community/followers/list/6a0acd09d287bdcbab076896"
    )

    response = requests.get(
        url,
        headers=get_headers(),
        timeout=30
    )

    json_dict = response.json()

    output = json_dict.get("data", [])

    for output_dict in output:

        post_list = [
            output_dict.get("follower_id"),
            output_dict.get("name"),
            output_dict.get("mobile"),
            output_dict.get("profileImage", "null")
        ]

        write_unique(post_list)

        print(post_list)

        if isunfollow:
            unfollow(output_dict["follower_id"])
            print("Unfollowed:", output_dict["follower_id"])

    return len(output)


def get_user(comm_id, page_id):
    response = requests.get(
        main_api1 + comm_url + comm_id + "&page=" + str(page_id),
        headers=get_headers(),
        timeout=30
    )

    json_dict = response.json()

    if not json_dict.get("success"):
        return 0

    posts = json_dict["data"]["posts"]

    for post in posts:
        follow(post["user_id"])

    return len(posts)
get_user("6a0d5a03a4f9e66563c207be",9)
save()