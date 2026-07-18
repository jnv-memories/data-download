import os
import requests
from openpyxl import Workbook, load_workbook
from auth import get_headers

main_api1 = "https://pw-api-gate.penpencil.co/v3/"
main_api2 = "https://api.penpencil.co"
EXCEL_FILE = "teacher_details.xlsx"
ID_FILE = "teacher_id.txt"
POINTER_FILE = "teacher_pointer.txt"
USER_FILE = "user_id.txt"

def unfollow(user):
    url = main_api1 + "community/followers/"
    payload = {"follow": False}
    requests.post(
        url + user,
        headers=get_headers(),
        json=payload,
        timeout=30
    )

def write_unique(post_list):
    user_id = str(post_list[0]).strip()
    if not os.path.exists(ID_FILE):
        open(ID_FILE, "w").close()
    with open(ID_FILE, "r", encoding="utf-8") as f:
        if user_id in {line.strip() for line in f}:
            return
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Teacher"
        ws.append([
            "Teacher ID",
            "Name",
            "Mobile No",
            "Profile Pic"
        ])
        wb.save(EXCEL_FILE)
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(post_list)
    wb.save(EXCEL_FILE)
    wb.close()
    with open(ID_FILE, "a", encoding="utf-8") as f:
        f.write(user_id + "\n")
def phone(user_id):
    if user_id not in ["6a0acd09d287bdcbab076896","691028ae9961d23389b2b7a2"]:
        url = main_api1 + "community/followers/list/" + str(user_id)
        response = requests.get(
            url,
            headers=get_headers(),
            timeout=30
        )
        json_dict = response.json()
        data = json_dict.get("data", [])
        if not data:
            print("No followers found.")
            return 0
        count = 0
        for output_dict in data:
            post_list = [
                output_dict.get("follower_id"),
                output_dict.get("name"),
                output_dict.get("mobile"),
                output_dict.get("profileImage", "null")
            ]
            write_unique(post_list)
            print(post_list)
            count += 1
    return count
def teacher():
    if not os.path.exists(USER_FILE):
        raise FileNotFoundError(
            "user_id.txt not found. Run save() first."
        )
    if not os.path.exists(POINTER_FILE):
        with open(POINTER_FILE, "w") as f:
            f.write("0")
    with open(POINTER_FILE, "r", encoding="utf-8") as f:
        pointer = int(f.read().strip() or "0")
    processed = 0
    with open(USER_FILE, "r", encoding="utf-8") as f:
        f.seek(pointer)
        for line in f:
            user_id = line.strip()
            if user_id:
                phone(user_id)
                processed += 1
        new_pointer = f.tell()
    with open(POINTER_FILE, "w", encoding="utf-8") as f:
        f.write(str(new_pointer))
    print(f"Processed {processed} users.")
    return processed

def teacher_details():
    url = main_api2+"/v3/batches/69897f0a4c12aeb013d4ea52/details"
    response = requests.get(
        url,
        headers=get_headers(),
        timeout=30
    )
    teacher_id = response["data"]["subjects"][1]["teacherIds"][0]["_id"]
    return teacher_id

if __name__ == "__main__":
    teacher()